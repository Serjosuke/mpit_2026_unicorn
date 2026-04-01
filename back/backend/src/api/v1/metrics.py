from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select

from src.api.deps import get_current_user, require_roles
from src.db.deps import DBSession
from src.models.certificate import Certificate
from src.models.course import Course
from src.models.course_lesson import CourseLesson
from src.models.course_module import CourseModule
from src.models.department import Department
from src.models.enrollment import Enrollment
from src.models.external_course_request import ExternalCourseRequest
from src.models.lesson_progress import LessonProgress
from src.models.review import Review
from src.models.user import User
from src.schemas.metrics import HRDashboardMetricsOut

router = APIRouter()


THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)
SUBTITLE_FONT = Font(size=11, bold=True)

STATUS_FILLS = {
    "critical": PatternFill("solid", fgColor="FDE9E7"),
    "warning": PatternFill("solid", fgColor="FFF4CC"),
    "ok": PatternFill("solid", fgColor="E2F0D9"),
    "completed": PatternFill("solid", fgColor="D9EAF7"),
    "not_started": PatternFill("solid", fgColor="EDEDED"),
}


def _style_header(ws, row_num: int = 1):
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_area(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def _autosize_columns(ws, min_width: int = 12, max_width: int = 35):
    for column_cells in ws.columns:
        length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > length:
                length = len(value)
        ws.column_dimensions[column_letter].width = max(min_width, min(length + 2, max_width))


def _write_kpi_card(ws, row: int, title: str, value, fill_color: str = "D9EAF7"):
    ws[f"A{row}"] = title
    ws[f"B{row}"] = value

    ws[f"A{row}"].font = SUBTITLE_FONT
    ws[f"B{row}"].font = Font(size=12, bold=True)

    ws[f"A{row}"].fill = PatternFill("solid", fgColor=fill_color)
    ws[f"B{row}"].fill = PatternFill("solid", fgColor=fill_color)

    ws[f"A{row}"].border = THIN_BORDER
    ws[f"B{row}"].border = THIN_BORDER

    ws[f"A{row}"].alignment = Alignment(vertical="center")
    ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")


def _role_scope_clause(current_user: User):
    if current_user.role == "manager":
        return User.manager_id == current_user.id
    return True


def _build_rows(db: DBSession, current_user: User, department_id: str | None = None):
    users_stmt = select(User).where(User.role == "employee", User.is_active == True)
    if current_user.role == "manager":
        users_stmt = users_stmt.where(User.manager_id == current_user.id)
    if department_id:
        users_stmt = users_stmt.where(User.department_id == department_id)

    employees = list(db.scalars(users_stmt.order_by(User.last_name, User.first_name)).all())
    rows = []

    for employee in employees:
        dep = db.get(Department, employee.department_id) if employee.department_id else None
        enrollments = list(
            db.scalars(select(Enrollment).where(Enrollment.user_id == employee.id)).all()
        )

        active = [e for e in enrollments if e.status in ["enrolled", "in_progress"]]
        completed = [e for e in enrollments if e.status == "completed"]

        course_rows = []
        progress_values = []
        problem_courses = 0
        max_sprint_lag = 0
        latest_activity_at = employee.created_at

        for enrollment in enrollments:
            course = db.get(Course, enrollment.course_id)
            if not course:
                continue

            lesson_ids = list(
                db.scalars(
                    select(CourseLesson.id)
                    .join(CourseModule, CourseLesson.module_id == CourseModule.id)
                    .where(CourseModule.course_id == course.id)
                ).all()
            )

            lesson_count = len(lesson_ids)
            done_count = 0

            if lesson_ids:
                done_count = len(
                    db.scalars(
                        select(LessonProgress).where(
                            LessonProgress.enrollment_id == enrollment.id,
                            LessonProgress.lesson_id.in_(lesson_ids),
                            LessonProgress.is_completed == True,
                        )
                    ).all()
                )

            progress = float(enrollment.progress_percent or 0)
            progress_values.append(progress)

            enrollment_updated_at = enrollment.updated_at or enrollment.created_at
            if enrollment_updated_at and (
                latest_activity_at is None or enrollment_updated_at > latest_activity_at
            ):
                latest_activity_at = enrollment_updated_at

            planned = (
                (enrollment.started_at or enrollment.created_at) + timedelta(days=max(30, lesson_count * 5))
                if enrollment
                else None
            )

            sprint_lag = 0
            if lesson_count:
                expected_ratio = min(
                    (
                        (datetime.now(timezone.utc) - (enrollment.started_at or enrollment.created_at)).days
                        / max(30, lesson_count * 5)
                    ),
                    1,
                )
                expected_done = int(expected_ratio * lesson_count)
                sprint_lag = max((expected_done - done_count + 1) // 2, 0)

            if enrollment.status == "completed":
                course_status = "Обучение завершено"
                course_status_group = "completed"
            elif progress == 0:
                course_status = "Не начал учёбу"
                course_status_group = "not_started"
            elif sprint_lag >= 3:
                course_status = "Отстаёт на 3+ спринта"
                course_status_group = "critical"
            elif sprint_lag >= 1:
                course_status = "Отстаёт на 1–2 спринта"
                course_status_group = "warning"
            else:
                course_status = "Успевает"
                course_status_group = "ok"

            if course_status_group in {"critical", "warning"}:
                problem_courses += 1

            max_sprint_lag = max(max_sprint_lag, sprint_lag)

            age_days_course = (
                (datetime.now(timezone.utc) - enrollment_updated_at).days
                if enrollment_updated_at
                else 999
            )
            if age_days_course == 0:
                course_activity_label = "Сегодня"
            elif age_days_course <= 7:
                course_activity_label = "На этой неделе"
            elif age_days_course <= 30:
                course_activity_label = "Больше недели назад"
            else:
                course_activity_label = "Больше месяца назад"

            course_rows.append(
                {
                    "course_title": course.title,
                    "progress_percent": round(progress, 2),
                    "last_activity": course_activity_label,
                    "last_activity_at": enrollment_updated_at.isoformat() if enrollment_updated_at else None,
                    "planned_completion": planned.date().isoformat() if planned else None,
                    "sprint_lag": sprint_lag,
                    "status": course_status,
                    "status_group": course_status_group,
                }
            )

        avg_progress = round(sum(progress_values) / len(progress_values), 2) if progress_values else 0

        age_days = (
            (datetime.now(timezone.utc) - latest_activity_at).days
            if latest_activity_at
            else 999
        )
        if age_days == 0:
            activity_label = "Сегодня"
        elif age_days <= 7:
            activity_label = "На этой неделе"
        elif age_days <= 30:
            activity_label = "Больше недели назад"
        else:
            activity_label = "Больше месяца назад"

        if not enrollments:
            employee_status = "Не назначены курсы"
            employee_status_group = "not_started"
        elif problem_courses > 0 and max_sprint_lag >= 3:
            employee_status = "Есть критично отстающие курсы"
            employee_status_group = "critical"
        elif problem_courses > 0:
            employee_status = "Есть курсы с риском"
            employee_status_group = "warning"
        elif len(active) == 0 and len(completed) > 0:
            employee_status = "Все активные курсы завершены"
            employee_status_group = "completed"
        else:
            employee_status = "Обучение в норме"
            employee_status_group = "ok"

        rows.append(
            {
                "user_id": str(employee.id),
                "employee_name": " ".join(
                    x for x in [employee.last_name, employee.first_name, employee.middle_name] if x
                ),
                "department_name": dep.name if dep else "Без отдела",
                "team_name": employee.team_name or "Без команды",
                "position_title": employee.position_title or employee.role,
                "manager_id": str(employee.manager_id) if employee.manager_id else None,
                "avg_progress_percent": avg_progress,
                "last_activity": activity_label,
                "last_activity_at": latest_activity_at.isoformat() if latest_activity_at else None,
                "max_sprint_lag": max_sprint_lag,
                "status": employee_status,
                "status_group": employee_status_group,
                "completed_courses": len(completed),
                "started_courses": len(enrollments),
                "active_courses": len(active),
                "problem_courses": problem_courses,
                "course_rows": course_rows,
            }
        )

    return rows


@router.get("/hr-dashboard", response_model=HRDashboardMetricsOut)
def hr_dashboard_metrics(db: DBSession, _: User = Depends(require_roles("admin", "hr"))):
    total_users = db.scalar(select(func.count()).select_from(User)) or 0
    total_employees = db.scalar(select(func.count()).select_from(User).where(User.role == "employee")) or 0
    total_trainers = db.scalar(select(func.count()).select_from(User).where(User.role == "trainer")) or 0
    total_courses = db.scalar(select(func.count()).select_from(Course)) or 0
    published_courses = db.scalar(select(func.count()).select_from(Course).where(Course.status == "published")) or 0
    total_enrollments = db.scalar(select(func.count()).select_from(Enrollment)) or 0
    completed_enrollments = db.scalar(select(func.count()).select_from(Enrollment).where(Enrollment.status == "completed")) or 0
    active_enrollments = db.scalar(select(func.count()).select_from(Enrollment).where(Enrollment.status.in_(["enrolled", "in_progress"]))) or 0
    completion_rate_percent = round((completed_enrollments / total_enrollments) * 100, 2) if total_enrollments else 0.0
    external_requests_total = db.scalar(select(func.count()).select_from(ExternalCourseRequest)) or 0
    external_requests_pending_manager = db.scalar(select(func.count()).select_from(ExternalCourseRequest).where(ExternalCourseRequest.status == "pending_manager_approval")) or 0
    external_requests_pending_hr = db.scalar(select(func.count()).select_from(ExternalCourseRequest).where(ExternalCourseRequest.status == "pending_hr_approval")) or 0
    external_requests_approved = db.scalar(select(func.count()).select_from(ExternalCourseRequest).where(ExternalCourseRequest.status == "approved")) or 0
    external_requests_rejected = db.scalar(select(func.count()).select_from(ExternalCourseRequest).where(ExternalCourseRequest.status.in_(["rejected_by_manager", "rejected_by_hr"]))) or 0
    certificates_total = db.scalar(select(func.count()).select_from(Certificate)) or 0
    reviews_total = db.scalar(select(func.count()).select_from(Review)) or 0
    average_review_rating = db.scalar(select(func.avg(Review.rating)).select_from(Review)) or 0
    average_review_rating = round(float(average_review_rating), 2)

    return HRDashboardMetricsOut(
        total_users=total_users,
        total_employees=total_employees,
        total_trainers=total_trainers,
        total_courses=total_courses,
        published_courses=published_courses,
        total_enrollments=total_enrollments,
        completed_enrollments=completed_enrollments,
        active_enrollments=active_enrollments,
        completion_rate_percent=completion_rate_percent,
        external_requests_total=external_requests_total,
        external_requests_pending_manager=external_requests_pending_manager,
        external_requests_pending_hr=external_requests_pending_hr,
        external_requests_approved=external_requests_approved,
        external_requests_rejected=external_requests_rejected,
        certificates_total=certificates_total,
        reviews_total=reviews_total,
        average_review_rating=average_review_rating,
    )


@router.get("/monitor")
def monitor_metrics(db: DBSession, current_user: User = Depends(require_roles("admin", "hr", "manager")), department_id: str | None = Query(default=None)):
    rows = _build_rows(db, current_user, department_id=department_id)
    summary = {
        "critical": len([r for r in rows if r["status_group"] == "critical"]),
        "warning": len([r for r in rows if r["status_group"] == "warning"]),
        "ok": len([r for r in rows if r["status_group"] == "ok"]),
        "completed": len([r for r in rows if r["status_group"] == "completed"]),
        "not_started": len([r for r in rows if r["status_group"] == "not_started"]),
    }
    department_cards = []
    departments = list(db.scalars(select(Department).order_by(Department.name)).all())
    for dep in departments:
        dep_rows = [r for r in rows if r["department_name"] == dep.name]
        if dep_rows:
            department_cards.append({"id": str(dep.id), "name": dep.name, "employees": len(dep_rows), "teams": len(set(r["team_name"] for r in dep_rows))})
    team_cards = []
    for team in sorted(set(r["team_name"] for r in rows)):
        team_rows = [r for r in rows if r["team_name"] == team]
        team_cards.append({"team_name": team, "employees": len(team_rows), "departments": sorted(set(r["department_name"] for r in team_rows))})
    return {"summary": summary, "departments": department_cards, "teams": team_cards, "rows": rows}


@router.get("/monitor-export")
def monitor_export(
    db: DBSession,
    current_user: User = Depends(require_roles("admin", "hr", "manager")),
    department_id: str | None = Query(default=None),
):
    rows = _build_rows(db, current_user, department_id=department_id)

    summary = {
        "critical": len([r for r in rows if r["status_group"] == "critical"]),
        "warning": len([r for r in rows if r["status_group"] == "warning"]),
        "ok": len([r for r in rows if r["status_group"] == "ok"]),
        "completed": len([r for r in rows if r["status_group"] == "completed"]),
        "not_started": len([r for r in rows if r["status_group"] == "not_started"]),
    }

    department_stats = defaultdict(
        lambda: {
            "employees": 0,
            "completed": 0,
            "in_progress": 0,
            "critical": 0,
            "warning": 0,
            "avg_progress_sum": 0.0,
        }
    )

    team_counter = Counter()

    for row in rows:
        dep = row["department_name"]
        team = row["team_name"]

        department_stats[dep]["employees"] += 1
        department_stats[dep]["avg_progress_sum"] += float(row["avg_progress_percent"] or 0)
        team_counter[(dep, team)] += 1

        if row["status_group"] == "completed":
            department_stats[dep]["completed"] += 1
        elif row["status_group"] in {"critical", "warning", "ok"}:
            department_stats[dep]["in_progress"] += 1

        if row["status_group"] == "critical":
            department_stats[dep]["critical"] += 1
        if row["status_group"] == "warning":
            department_stats[dep]["warning"] += 1

    wb = Workbook()

    ws_dashboard = wb.active
    ws_dashboard.title = "Дашборд"

    ws_dashboard["A1"] = "HR-отчёт по обучению сотрудников"
    ws_dashboard["A1"].font = TITLE_FONT

    ws_dashboard["A2"] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_dashboard["A2"].font = Font(italic=True, color="666666")

    _write_kpi_card(ws_dashboard, 4, "Всего сотрудников", len(rows), "D9EAF7")
    _write_kpi_card(ws_dashboard, 5, "Все активные курсы завершены", summary["completed"], "DDEBF7")
    _write_kpi_card(ws_dashboard, 6, "Обучение в норме", summary["ok"], "E2F0D9")
    _write_kpi_card(ws_dashboard, 7, "Есть курсы с риском", summary["warning"], "FFF4CC")
    _write_kpi_card(ws_dashboard, 8, "Критично отстают", summary["critical"], "FDE9E7")
    _write_kpi_card(ws_dashboard, 9, "Не назначены курсы", summary["not_started"], "EDEDED")

    avg_progress = (
        round(sum(float(r["avg_progress_percent"] or 0) for r in rows) / len(rows), 2)
        if rows
        else 0
    )
    _write_kpi_card(ws_dashboard, 10, "Средний прогресс, %", avg_progress, "EADCF8")

    ws_dashboard["D4"] = "Распределение по статусам"
    ws_dashboard["D4"].font = SUBTITLE_FONT

    ws_dashboard["D5"] = "Статус"
    ws_dashboard["E5"] = "Кол-во"

    for cell in ws_dashboard[5][3:5]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    status_rows = [
        ("Критично отстают", summary["critical"], "critical"),
        ("Есть курсы с риском", summary["warning"], "warning"),
        ("Обучение в норме", summary["ok"], "ok"),
        ("Все активные курсы завершены", summary["completed"], "completed"),
        ("Не назначены курсы", summary["not_started"], "not_started"),
    ]

    current_row = 6
    for label, value, status_group in status_rows:
        ws_dashboard[f"D{current_row}"] = label
        ws_dashboard[f"E{current_row}"] = value
        ws_dashboard[f"D{current_row}"].fill = STATUS_FILLS[status_group]
        ws_dashboard[f"E{current_row}"].fill = STATUS_FILLS[status_group]
        ws_dashboard[f"D{current_row}"].border = THIN_BORDER
        ws_dashboard[f"E{current_row}"].border = THIN_BORDER
        current_row += 1

    ws_dashboard["D13"] = "Отделы"
    ws_dashboard["D13"].font = SUBTITLE_FONT
    ws_dashboard["D14"] = "Отдел"
    ws_dashboard["E14"] = "Сотрудников"
    ws_dashboard["F14"] = "Критично"
    ws_dashboard["G14"] = "Средний прогресс %"

    for cell in ws_dashboard[14][3:7]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    dep_row_idx = 15
    sorted_departments = sorted(
        department_stats.items(),
        key=lambda x: x[1]["employees"],
        reverse=True,
    )

    for dep_name, stats in sorted_departments:
        avg_dep_progress = (
            round(stats["avg_progress_sum"] / stats["employees"], 2)
            if stats["employees"]
            else 0
        )
        ws_dashboard[f"D{dep_row_idx}"] = dep_name
        ws_dashboard[f"E{dep_row_idx}"] = stats["employees"]
        ws_dashboard[f"F{dep_row_idx}"] = stats["critical"]
        ws_dashboard[f"G{dep_row_idx}"] = avg_dep_progress
        for col in ["D", "E", "F", "G"]:
            ws_dashboard[f"{col}{dep_row_idx}"].border = THIN_BORDER
        dep_row_idx += 1

    ws_dashboard.freeze_panes = "A4"
    _autosize_columns(ws_dashboard)

    ws_employees = wb.create_sheet(title="Сотрудники")

    employee_headers = [
        "Сотрудник",
        "Отдел",
        "Команда",
        "Роль в команде",
        "Пройдено курсов",
        "Активных курсов",
        "Всего курсов",
        "Проблемных курсов",
        "Средний прогресс %",
        "Последняя активность",
        "Дата последней активности",
        "Макс. отставание (спринты)",
        "Статус",
        "Группа статуса",
    ]

    ws_employees.append(employee_headers)
    _style_header(ws_employees, 1)

    for row in rows:
        ws_employees.append(
            [
                row["employee_name"],
                row["department_name"],
                row["team_name"],
                row["position_title"],
                row["completed_courses"],
                row["active_courses"],
                row["started_courses"],
                row["problem_courses"],
                row["avg_progress_percent"],
                row["last_activity"],
                row["last_activity_at"],
                row["max_sprint_lag"],
                row["status"],
                row["status_group"],
            ]
        )

    if rows:
        for idx in range(2, len(rows) + 2):
            status_group = ws_employees[f"N{idx}"].value
            fill = STATUS_FILLS.get(status_group)
            if fill:
                for col in range(1, len(employee_headers) + 1):
                    ws_employees.cell(row=idx, column=col).fill = fill

    _style_data_area(ws_employees, 2, max(2, len(rows) + 1), 1, len(employee_headers))
    ws_employees.auto_filter.ref = (
        f"A1:{get_column_letter(len(employee_headers))}{max(1, len(rows) + 1)}"
    )
    ws_employees.freeze_panes = "A2"
    _autosize_columns(ws_employees)

    ws_courses = wb.create_sheet(title="Курсы сотрудников")

    course_headers = [
        "Сотрудник",
        "Отдел",
        "Команда",
        "Роль в команде",
        "Курс",
        "Прогресс %",
        "Последняя активность",
        "Дата последней активности",
        "Плановое завершение",
        "Отставание (спринты)",
        "Статус",
        "Группа статуса",
    ]

    ws_courses.append(course_headers)
    _style_header(ws_courses, 1)

    for row in rows:
        for course_row in row["course_rows"]:
            ws_courses.append(
                [
                    row["employee_name"],
                    row["department_name"],
                    row["team_name"],
                    row["position_title"],
                    course_row["course_title"],
                    course_row["progress_percent"],
                    course_row["last_activity"],
                    course_row["last_activity_at"],
                    course_row["planned_completion"],
                    course_row["sprint_lag"],
                    course_row["status"],
                    course_row["status_group"],
                ]
            )

    if ws_courses.max_row > 1:
        for idx in range(2, ws_courses.max_row + 1):
            status_group = ws_courses[f"L{idx}"].value
            fill = STATUS_FILLS.get(status_group)
            if fill:
                for col in range(1, len(course_headers) + 1):
                    ws_courses.cell(row=idx, column=col).fill = fill

    _style_data_area(ws_courses, 2, max(2, ws_courses.max_row), 1, len(course_headers))
    ws_courses.auto_filter.ref = f"A1:{get_column_letter(len(course_headers))}{max(1, ws_courses.max_row)}"
    ws_courses.freeze_panes = "A2"
    _autosize_columns(ws_courses)

    ws_departments = wb.create_sheet(title="По отделам")

    dep_headers = [
        "Отдел",
        "Сотрудников",
        "Команд",
        "Все активные курсы завершены",
        "В процессе / в норме",
        "Критично отстают",
        "Есть риски",
        "Средний прогресс %",
    ]
    ws_departments.append(dep_headers)
    _style_header(ws_departments, 1)

    for dep_name, stats in sorted_departments:
        teams_count = len([team for (dep, team), _count in team_counter.items() if dep == dep_name])
        avg_dep_progress = (
            round(stats["avg_progress_sum"] / stats["employees"], 2)
            if stats["employees"]
            else 0
        )
        ws_departments.append(
            [
                dep_name,
                stats["employees"],
                teams_count,
                stats["completed"],
                stats["in_progress"],
                stats["critical"],
                stats["warning"],
                avg_dep_progress,
            ]
        )

    _style_data_area(ws_departments, 2, max(2, len(sorted_departments) + 1), 1, len(dep_headers))
    ws_departments.auto_filter.ref = (
        f"A1:{get_column_letter(len(dep_headers))}{max(1, len(sorted_departments) + 1)}"
    )
    ws_departments.freeze_panes = "A2"
    _autosize_columns(ws_departments)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="alrosa-monitor.xlsx"'},
    )


@router.get("/home")
def home_metrics(db: DBSession, current_user: User = Depends(get_current_user)):
    enrollments = list(db.scalars(select(Enrollment).where(Enrollment.user_id == current_user.id).order_by(Enrollment.updated_at.desc())).all())
    certificates = list(db.scalars(select(Certificate).where(Certificate.user_id == current_user.id)).all())
    active = [e for e in enrollments if e.status in ["enrolled", "in_progress"]]
    urgent = None
    for e in active:
        course = db.get(Course, e.course_id)
        lesson_ids = list(db.scalars(select(CourseLesson.id).join(CourseModule, CourseLesson.module_id == CourseModule.id).where(CourseModule.course_id == e.course_id)).all())
        progress_map = {r.lesson_id: r for r in db.scalars(select(LessonProgress).where(LessonProgress.enrollment_id == e.id)).all()}
        modules = list(db.scalars(select(CourseModule).where(CourseModule.course_id == e.course_id).order_by(CourseModule.order_index)).all())
        base_date = e.started_at or e.created_at
        for module_index, module in enumerate(modules):
            lessons = list(db.scalars(select(CourseLesson).where(CourseLesson.module_id == module.id).order_by(CourseLesson.order_index)).all())
            for lesson in lessons:
                pr = progress_map.get(lesson.id)
                if pr and pr.is_completed:
                    continue
                due_at = base_date + timedelta(days=(module_index * 7) + lesson.order_index * 2)
                status = 'danger' if due_at < datetime.now(timezone.utc) else ('warning' if due_at - datetime.now(timezone.utc) <= timedelta(hours=24) else 'normal')
                candidate = {
                    'course_title': course.title if course else '—',
                    'lesson_title': lesson.title,
                    'due_at': due_at.isoformat(),
                    'deadline_status': status,
                    'enrollment_id': str(e.id),
                    'course_id': str(e.course_id),
                }
                if urgent is None or due_at < datetime.fromisoformat(urgent['due_at']):
                    urgent = candidate
    return {
        'active_courses': len(active),
        'completed_courses': len([e for e in enrollments if e.status == 'completed']),
        'certificates': len(certificates),
        'urgent_lesson': urgent,
    }
